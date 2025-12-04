import os
import time
import glob
import random
import json
from urllib.parse import urljoin
import os  # estado persistente y paths
import json  # leer/escribir state.json
import time  # timestamp de última ejecución
import random  # esperas humanizadas

from camoufox import Camoufox
from rich.console import Console
from openpyxl import Workbook, load_workbook
from threading import Thread, Event
from queue import Queue, Empty

console = Console(force_terminal=True, color_system="truecolor", markup=False, soft_wrap=False)

# Ruta base para perfiles; crearemos un subdirectorio por sesión para evitar arrastre tras WAF
USER_DATA_DIR_BASE = (
    '/Users/gonzalocamunez/Documents/Python Windsurf/'
    'scraper_*********/user_data_dir'
)

# === NUEVO SISTEMA DIARIO EN 2 FASES (Estado persistente) ===
# Persistimos estado de ejecución diario para implementar la lógica solicitada:
#  - ids_conocidos: set/listado de id_anuncio ya procesados (scrapeados)
#  - ultimo_pivote_fase1: id_anuncio donde Fase 1 decidió parar por racha de conocidos
#  - ultimo_procesado_fase2: id_anuncio más antiguo procesado en la Fase 2 (marcador de reanudación)
#  - fecha_ultimo_proceso: timestamp de última ejecución (solo informativo)
STATE_FILE = os.path.join(os.path.dirname(__file__), 'state.json')  # Ruta de estado

# === COOKIE STORE Y ESTRATEGIA DE FASE 2 ===
# Persistencia manual de cookies para contextos efímeros
COOKIE_STORE_PATH = os.path.join(os.path.dirname(__file__), 'cookies_*********.json')
# Configuración de Fase 2 (sticky con sesstime opcional)
PHASE2_USE_STICKY = True
PHASE2_SESSTIME_SEC = 600  # 10 minutos

# === Rotación periódica (modo eficiente) ===
# Umbrales base; se aplicará jitter por sesión (±)
ROTATE_TIME_MIN = 15
ROTATE_TIME_JITTER = 3
ROTATE_PAGES = 16
ROTATE_PAGES_JITTER = 3

# URLs base relevantes (orden reciente primero). La ascendente se simula con estrategia de páginas.
WEB_LISTADO_DESC = 'https://www.*********.com/venta-viviendas/madrid-madrid/?ordenado-por=fecha-publicacion-desc'
WEB_LISTADO_ASC = 'https://www.*********.com/venta-viviendas/madrid-madrid/?ordenado-por=fecha-publicacion-asc'

def load_cookies_from_file(path: str = COOKIE_STORE_PATH):
    """Carga cookies Playwright-compatibles desde disco."""
    try:
        if not os.path.exists(path):
            return []
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, dict) and 'cookies' in data:
            return data.get('cookies') or []
        if isinstance(data, list):
            return data
        return []
    except Exception as e:
        try:
            log_warn(f"No se pudieron cargar cookies desde {path}: {e}")
        except Exception:
            pass
        return []

def save_cookies(page, path: str = COOKIE_STORE_PATH):
    """Guarda cookies actuales del contexto en disco."""
    try:
        cookies = page.context.cookies()
        payload = {"cookies": cookies, "saved": time.time()}
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2)
        log_sys(f"Cookies guardadas en {path} ({len(cookies)} items)")
    except Exception as e:
        log_warn(f"No se pudieron guardar cookies en {path}: {e}")

def try_inject_cookies(page, path: str = COOKIE_STORE_PATH):
    """Intenta inyectar cookies guardadas en el contexto actual antes de navegar."""
    try:
        cookies = load_cookies_from_file(path)
        if not cookies:
            return False
        # Normalizar estructura mínima
        norm = []
        for c in cookies:
            try:
                cc = {
                    'name': c['name'],
                    'value': c['value'],
                    'domain': c.get('domain') or '.*********.com',
                    'path': c.get('path') or '/',
                }
                if 'expires' in c and c['expires'] is not None:
                    cc['expires'] = c['expires']
                if 'httpOnly' in c:
                    cc['httpOnly'] = c['httpOnly']
                if 'secure' in c:
                    cc['secure'] = c['secure']
                if 'sameSite' in c and c['sameSite']:
                    cc['sameSite'] = c['sameSite']
                norm.append(cc)
            except Exception:
                continue
        if not norm:
            return False
        page.context.add_cookies(norm)
        log_sys(f"Inyectadas {len(norm)} cookies desde {path}")
        return True
    except Exception as e:
        log_warn(f"No se pudieron inyectar cookies desde {path}: {e}")
        return False

def load_persistent_state(path: str = STATE_FILE) -> dict:
    """Carga el estado persistente del scraper (o estructura por defecto).

    Devuelve dict con claves compatibles anteriores y nuevas:
    - ids_conocidos (legacy)
    - ultimo_pivote_fase1, ultimo_procesado_fase2 (legacy)
    - fecha_ultimo_proceso (legacy)
    - pivots: {recientes: {pivot_id, ts}, antiguos: {pivot_id, ts}}
    - siguiente_ejecucion: 'recientes' | 'antiguos'
    - consecutive_200: int
    - retry_queue: [ {id, url, reason, attempts, next_try_at} ]
    - counts: {'recientes': 0, 'antiguos': 0, 'total': 0}
    """
    state = {
        'ids_conocidos': [],
        'ultimo_pivote_fase1': None,
        'ultimo_procesado_fase2': None,
        'fecha_ultimo_proceso': None,
        'pivots': {
            'recientes': {'pivot_id': None, 'ts': None},
            'antiguos': {'pivot_id': None, 'ts': None},
        },
        'siguiente_ejecucion': 'recientes',
        'consecutive_200': 0,
        'retry_queue': [],
        'counts': {'recientes': 0, 'antiguos': 0, 'total': 0},
    }
    try:
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # merge con defaults para robustez
            state.update({k: data.get(k, state[k]) for k in state.keys()})
    except Exception as e:
        log_warn(f"No se pudo cargar estado desde {path}: {e}")
    # Normalizar tipos
    try:
        ids = state.get('ids_conocidos') or []
        state['ids_conocidos'] = [str(x) for x in ids if x]
    except Exception:
        state['ids_conocidos'] = []
    return state


def _init_rotation_limits():
    """Genera umbrales de rotación con jitter para la sesión actual.

    Returns: (time_limit_sec, pages_limit)
    """
    try:
        t = (ROTATE_TIME_MIN + random.randint(-ROTATE_TIME_JITTER, ROTATE_TIME_JITTER))
        if t < 5:
            t = 5
    except Exception:
        t = ROTATE_TIME_MIN
    try:
        p = (ROTATE_PAGES + random.randint(-ROTATE_PAGES_JITTER, ROTATE_PAGES_JITTER))
        if p < 4:
            p = 4
    except Exception:
        p = ROTATE_PAGES
    return t * 60, p


def initial_warm_up(sessid: str):
    """ETAPA 1: Warm‑up inicial para obtener cookie válida y estabilizar el proxy.

    - Sticky session
    - Inyección/guardado de cookies
    """
    global COOKIE_STORE_PATH
    # Determinar si venimos de WAF alto para entrar en "modo seguro"
    try:
        __st = load_persistent_state()
        __safe_mode = bool(__st.get('last_waf_high'))
    except Exception:
        __safe_mode = False

    with Camoufox(
        headless=False,
        main_world_eval=False,
        persistent_context=False,
        enable_cache=False,
        block_webrtc=True,
        block_webgl=True,
        block_images=False,
        i_know_what_im_doing=True,
        os="macos",
        fonts=get_mac_fonts(sessid, 12),
        geoip=True,
        humanize=True,
        proxy={
            'server': STICKY_PROXY_SERVER,
            'username': f"{PROXY_USERNAME}-sessid-{sessid}",
            'password': PROXY_PASSWORD,
        }
    ) as browser:
        page = browser.new_page()
        page.set_default_navigation_timeout(90000)
        page.set_default_timeout(90000)
        # No inyectar cookies en warm_up para evitar reutilizar posibles cookies "flageadas"
        waf_high = False
        status = navigate_and_stabilize(page, "https://www.*********.com/", tag="warmup_home")
        risk = classify_waf_risk(status)
        if risk in ("MUY ALTO", "ALTO"):
            log_waf(f"{risk} en warm_up; continuar con cautela")
            # Marcar WAF alto en estado para que etapas posteriores no inyecten/guarden cookies
            try:
                _st = load_persistent_state()
                _st['last_waf_high'] = True
                save_persistent_state(_st)
                waf_high = True
            except Exception:
                pass
        gentle_scroll(page)
        safe_wait_page(page, random.randint(400, 800))
        # Guardar cookies solo si 200 estable y no venimos de WAF alto
        if (str(status).isdigit() and int(status) == 200) and not waf_high:
            try:
                _st = load_persistent_state()
                _st['last_waf_high'] = False
                save_persistent_state(_st)
            except Exception:
                pass
            try:
                save_cookies(page, COOKIE_STORE_PATH)
            except Exception:
                pass


def _enqueue_retry(state: dict, lid: str, url: str, reason: str, attempts: int, delay_sec: int):
    """Añade un item a la cola de reintentos con next_try_at.

    Nota: a partir de ND 5/9/2025, se permite incluir 'source' (recientes|antiguos)
    en el dict de entrada si está disponible.
    """
    try:
        next_ts = time.time() + delay_sec
        entry = {
            'id': lid,
            'url': url,
            'reason': reason,
            'attempts': attempts,
            'next_try_at': next_ts,
        }
        rq = state.get('retry_queue') or []
        rq.append(entry)
        state['retry_queue'] = rq
        save_persistent_state(state)
    except Exception as e:
        log_warn(f"No se pudo encolar retry: {e}")


def _drain_ready_retries_to_queue(state: dict, q: Queue, known_ids: set):
    """Mueve a la cola los reintentos cuyo next_try_at ya venció y no son conocidos."""
    try:
        now = time.time()
        remain = []
        moved = 0
        for it in (state.get('retry_queue') or []):
            lid = str(it.get('id')) if it.get('id') is not None else None
            url = it.get('url')
            src = it.get('source')
            if not lid or not url:
                continue
            if lid in known_ids:
                continue
            if it.get('next_try_at') and it['next_try_at'] <= now:
                try:
                    if src:
                        q.put_nowait((lid, url, src))
                    else:
                        q.put_nowait((lid, url))
                    moved += 1
                except Exception:
                    remain.append(it)
            else:
                remain.append(it)
        if moved:
            log_sys(f"Reinsertados {moved} items desde retry_queue")
        state['retry_queue'] = remain
        if moved:
            save_persistent_state(state)
    except Exception as e:
        log_warn(f"Error drenando retry_queue: {e}")


def stage2_producer(sessid: str, q: Queue, stop_event: Event, state: dict, known_ids: set,
                    max_pages: int = 100, consecutivos_stop: int = 500):
    """ETAPA 2: Productor de IDs (recientes o antiguos según pivote/siguiente_ejecucion).

    Inserta en la cola FIFO cada (id, url) nuevo.
    Persiste pivotes cuando encuentra el primer conocido.
    """
    direction = (state.get('siguiente_ejecucion') or 'recientes')
    if SHOW_NAV:
        log_sys(f"[STAGE2] Inicio productor en dirección={direction}")
    # Modo seguro si el último ciclo terminó con WAF alto
    __safe_mode = bool(state.get('last_waf_high'))

    with Camoufox(
        headless=True,
        main_world_eval=False,
        persistent_context=False,
        enable_cache=False,
        block_webrtc=True,
        block_webgl=True,
        block_images=False,
        i_know_what_im_doing=True,
        os="macos",
        fonts=get_mac_fonts(sessid, 12),
        geoip=True,
        humanize=True,
        proxy={
            'server': STICKY_PROXY_SERVER,
            'username': f"{PROXY_USERNAME}-sessid-{sessid}",
            'password': PROXY_PASSWORD,
        }
    ) as browser:
        page = browser.new_page()
        page.set_default_navigation_timeout(90000)
        page.set_default_timeout(90000)
        # Evitar reinyección si el ciclo previo terminó por WAF alto
        if not (state.get('last_waf_high') is True):
            try_inject_cookies(page, COOKIE_STORE_PATH)

        # Navegar a la lista según dirección
        if direction == 'recientes':
            _goto_listado_desc(page)
        else:
            _goto_listado_asc(page)

        # ========= Nueva lógica basada en pivotes =========
        # 1) Colectar una lista plana de (id, url) a lo largo de páginas (hasta max_pages)
        all_items = []  # [(id, url)] en orden visual absoluto
        page_idx = 1
        while page_idx <= max_pages and not stop_event.is_set():
            ids_urls = _extract_ids_from_current_listado(page)
            if ids_urls:
                all_items.extend(ids_urls)
            next_url = get_next_page_url(page)
            if not next_url:
                break
            status = navigate_and_stabilize(page, next_url, tag='s2_next', wait_networkidle=False)
            risk = classify_waf_risk(status)
            if risk in ("MUY ALTO", "ALTO"):
                # Non-stop: marcar reinicio inmediato y modo seguro, persistir y salir
                log_waf(f"[STAGE2] {risk} navegando listado; NEED_RESTART")
                try:
                    state['last_waf_high'] = True
                except Exception:
                    pass
                try:
                    save_persistent_state(state)
                except Exception:
                    pass
                try:
                    global NEED_RESTART
                    NEED_RESTART = True
                except Exception:
                    pass
                return
            gentle_scroll(page)
            safe_wait_page(page, random.randint(1600, 3200))
            page_idx += 1

        # 2) Determinar índice de arranque según reglas
        key = 'recientes' if direction == 'recientes' else 'antiguos'
        prev = (state.get('pivots') or {}).get(key, {}) or {}
        prev_pivot_id = str(prev.get('pivot_id')) if prev.get('pivot_id') is not None else None
        prev_pivot_n = int(prev.get('pivot_n') or 0)
        include_prev = bool(prev_pivot_id and (prev_pivot_id not in known_ids))

        start_idx = 1  # 1-based
        # Auxiliares para persistencia al final
        last_enq_idx = 0
        last_enq_id = None

        if direction == 'recientes':
            # Caso sin pivote previo (Paso 1): producir desconocidos desde el inicio
            if prev_pivot_n <= 0:
                start_idx = 1
            else:
                # Pasos 3–5: encontrar primer conocido K
                first_known_pos = None
                for i, (lid, _) in enumerate(all_items, start=1):
                    if lid in known_ids:
                        first_known_pos = i
                        break
                if first_known_pos is None:
                    # No hay conocido visible → producir desde inicio
                    start_idx = 1
                else:
                    # Contar conocidos consecutivos tras K
                    cnt_known_run = 0
                    for j in range(first_known_pos + 1, len(all_items) + 1):
                        if all_items[j - 1][0] in known_ids:
                            cnt_known_run += 1
                        else:
                            break
                    delta = max(prev_pivot_n - first_known_pos, 0)
                    delta_adj = max(delta - cnt_known_run, 0)
                    pos_target = first_known_pos + cnt_known_run + delta_adj
                    # SNAP si aparece el pivot_id previo
                    snap_pos = None
                    if prev_pivot_id:
                        for i, (lid, _) in enumerate(all_items, start=1):
                            if lid == prev_pivot_id:
                                snap_pos = i
                                break
                    if snap_pos is not None:
                        start_idx = snap_pos if include_prev else (snap_pos + 1)
                    else:
                        start_idx = max(1, min(pos_target, len(all_items)))

        else:  # antiguos
            if prev_pivot_n <= 0:
                start_idx = 1
            else:
                # Caer en n; ajustar por eliminaciones
                n = min(prev_pivot_n, len(all_items)) if len(all_items) > 0 else 1
                if n <= 0:
                    n = 1
                lid_n = all_items[n - 1][0] if all_items else None
                if prev_pivot_id and lid_n == prev_pivot_id:
                    start_idx = n if include_prev else (n + 1)
                else:
                    if lid_n and (lid_n in known_ids):
                        start_idx = n + 1
                    else:
                        # Retroceder hasta el primer conocido anterior
                        last_known_pos = None
                        for i in range(n - 1, 0, -1):
                            if all_items[i - 1][0] in known_ids:
                                last_known_pos = i
                                break
                        start_idx = (last_known_pos + 1) if last_known_pos else 1

        # 3) Encolar desde start_idx sólo los desconocidos
        for idx in range(start_idx, len(all_items) + 1):
            lid, lurl = all_items[idx - 1]
            if lid in known_ids:
                continue
            try:
                q.put((lid, lurl, direction))
                last_enq_idx = idx
                last_enq_id = lid
            except Exception as e:
                log_warn(f"No se pudo encolar {lid}: {e}")

        # 4) Persistir nuevo pivote_n y pivot_id según donde paramos
        try:
            if last_enq_idx > 0 and last_enq_idx <= len(all_items):
                state.setdefault('pivots', {})
                state['pivots'][key] = {
                    'pivot_n': last_enq_idx,
                    'pivot_id': last_enq_id,
                    'ts': time.time(),
                }
            # Alternar dirección para la próxima ejecución
            state['siguiente_ejecucion'] = 'antiguos' if direction == 'recientes' else 'recientes'
            save_persistent_state(state)
        except Exception:
            pass

        try:
            save_cookies(page, COOKIE_STORE_PATH)
        except Exception:
            pass
    # Señal de cierre para consumidor
    try:
        q.put_nowait(None)
    except Exception:
        pass
    stop_event.set()


def stage3_consumer(sessid: str, q: Queue, stop_event: Event, state: dict, known_ids: set,
                    time_limit_sec: int, pages_limit: int):
    """ETAPA 3: Consumidor en orden con backoff y cola diferida.

    - Visita `web_anuncio` en sticky.
    - Aplica backoff con jitter y motivos.
    - Aplica rotación periódica por tiempo/páginas señalizando NEED_RESTART.
    """
    global NEED_RESTART
    start_ts = time.time()
    last_rotate_ts = start_ts
    pages_since_rotate = 0
    local_attempts = {}

    with Camoufox(
        headless=True,
        main_world_eval=False,
        persistent_context=False,
        enable_cache=False,
        block_webrtc=True,
        block_webgl=True,
        block_images=False,
        i_know_what_im_doing=True,
        os="macos",
        fonts=get_mac_fonts(sessid, 12),
        geoip=True,
        humanize=True,
        proxy={
            'server': STICKY_PROXY_SERVER,
            'username': f"{PROXY_USERNAME}-sessid-{sessid}",
            'password': PROXY_PASSWORD,
        }
    ) as browser:
        page = browser.new_page()
        page.set_default_navigation_timeout(90000)
        page.set_default_timeout(90000)
        # Evitar reinyección si el ciclo previo terminó por WAF alto
        if not (state.get('last_waf_high') is True):
            try_inject_cookies(page, COOKIE_STORE_PATH)

        idx = 0
        # Micro-bursts en consumidor: breve think-time cada cierto número de anuncios
        next_consumer_burst = random.randint(9, 14)
        while True:
            # Reinsertar reintentos listos
            _drain_ready_retries_to_queue(state, q, known_ids)
            try:
                item = q.get(timeout=2)
            except Empty:
                if stop_event.is_set():
                    break
                else:
                    continue
            if item is None:
                break
            # Soportar tuplas (lid, url) o (lid, url, source)
            try:
                if isinstance(item, (tuple, list)) and len(item) == 3:
                    lid, url, source = item
                else:
                    lid, url = item
                    source = None
            except Exception:
                # fallback robusto
                try:
                    lid, url = item
                except Exception:
                    continue
                source = None
            if lid in known_ids:
                continue
            idx += 1
            if SHOW_STEPS:
                log_sys(f"[STAGE3] Visitando {idx}: {lid}")
            # Think-time muy corto para romper ritmos perfectos cada cierto número
            if (idx % next_consumer_burst) == 0:
                pause = random.randint(120, 300)
                if SHOW_BURSTS:
                    log_sys(f"[BURST] consumer micro-think ~{pause}ms antes de visitar")
                safe_wait_page(page, pause)
                next_consumer_burst = random.randint(9, 14)
            status = navigate_and_stabilize(page, url, tag='web_anuncio')
            risk = classify_waf_risk(status)

            # Rotación forzada si MUY_ALTO/ALTO y política de NEED_RESTART
            if risk in ("MUY ALTO", "ALTO"):
                # Registrar reintento
                attempts = int(local_attempts.get(lid, 0)) + 1
                local_attempts[lid] = attempts
                base = 6
                delay = int(base * (2 ** (attempts - 1)))
                jitter = int(delay * 0.25)
                delay += random.randint(-jitter, jitter)
                if attempts <= 3:
                    reason = f"HTTP_{status}" if status else "RISK_HIGH"
                    # incluir la fuente si se conoce
                    try:
                        if source:
                            entry = {
                                'id': lid, 'url': url, 'reason': reason,
                                'attempts': attempts, 'next_try_at': time.time() + max(delay, 8),
                                'source': source
                            }
                            rq = state.get('retry_queue') or []
                            rq.append(entry)
                            state['retry_queue'] = rq
                            save_persistent_state(state)
                        else:
                            _enqueue_retry(state, lid, url, reason, attempts, max(delay, 8))
                    except Exception:
                        _enqueue_retry(state, lid, url, reason, attempts, max(delay, 8))
                else:
                    log_waf(f"[STAGE3] {lid} agotó reintentos")
                # Señalizar reinicio para rotar seed
                NEED_RESTART = True
                # Avisar al productor para salida inmediata
                try:
                    stop_event.set()
                except Exception:
                    pass
                # Marcar en estado que hubo WAF alto para evitar reusar cookies en el siguiente arranque
                try:
                    state['last_waf_high'] = True
                except Exception:
                    pass
                save_persistent_state(state)
                break

            # Éxito: parsear y guardar
            try:
                humanize_anuncio_before_parse(page)
                data = parse_anuncio(page)
                if data and data.get('id_anuncio'):
                    xlsx_path = os.path.join(os.path.dirname(__file__), 'scraper.xlsx')
                    append_row_xlsx(xlsx_path, data)
                    known_ids.add(lid)
                    # Actualizar contadores por fuente
                    try:
                        counts = state.get('counts') or {'recientes': 0, 'antiguos': 0, 'total': 0}
                        if source == 'recientes':
                            counts['recientes'] = int(counts.get('recientes', 0)) + 1
                        elif source == 'antiguos':
                            counts['antiguos'] = int(counts.get('antiguos', 0)) + 1
                        counts['total'] = int(counts.get('total', 0)) + 1
                        state['counts'] = counts
                        # Log de progreso cada 10 items (opcional)
                        if SHOW_PROGRESS:
                            try:
                                if (counts.get('total', 0) % 10) == 0:
                                    log_sys(f"[PROGRESS] Recientes={counts.get('recientes',0)} · Antiguos={counts.get('antiguos',0)} · Total={counts.get('total',0)}")
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # consecutive_200
                    try:
                        if int(status) == 200:
                            # Incrementar racha de 200; salir de safe mode solo con >=5 consecutivos
                            c200 = int(state.get('consecutive_200', 0)) + 1
                            state['consecutive_200'] = c200
                            if c200 >= 5 and bool(state.get('last_waf_high')):
                                state['last_waf_high'] = False
                                # Reactivar rotación periódica en caliente
                                try:
                                    time_limit_sec, pages_limit = _init_rotation_limits()
                                    last_rotate_ts = time.time()
                                    pages_since_rotate = 0
                                    log_sys("[SAFE→EFFICIENT] ≥5 200: salimos de modo seguro y reactivamos rotación periódica")
                                except Exception:
                                    log_sys("[SAFE→EFFICIENT] ≥5 200: salimos de modo seguro")
                        else:
                            state['consecutive_200'] = 0
                    except Exception:
                        state['consecutive_200'] = 0
                    save_persistent_state(state)
                else:
                    # contenido vacío → reintento
                    attempts = int(local_attempts.get(lid, 0)) + 1
                    local_attempts[lid] = attempts
                    base = 6
                    delay = int(base * (2 ** (attempts - 1)))
                    jitter = int(delay * 0.25)
                    delay += random.randint(-jitter, jitter)
                    if attempts <= 3:
                        try:
                            if source:
                                entry = {
                                    'id': lid, 'url': url, 'reason': 'EMPTY_CONTENT',
                                    'attempts': attempts, 'next_try_at': time.time() + max(delay, 8),
                                    'source': source
                                }
                                rq = state.get('retry_queue') or []
                                rq.append(entry)
                                state['retry_queue'] = rq
                                save_persistent_state(state)
                            else:
                                _enqueue_retry(state, lid, url, 'EMPTY_CONTENT', attempts, max(delay, 8))
                        except Exception:
                            _enqueue_retry(state, lid, url, 'EMPTY_CONTENT', attempts, max(delay, 8))
                    else:
                        log_warn(f"[STAGE3] {lid} marcado como failed por contenido vacío")
            except Exception as e:
                log_warn(f"[STAGE3] Error extrayendo {lid}: {e}")

            # Scroll y pausa breve
            gentle_scroll(page)
            wait_ms = random.randint(1200, 2200)
            log_sys(f"[RHYTHM] visita_anuncio pausa ~{wait_ms}ms")
            safe_wait_page(page, wait_ms)

            # Rotación periódica por páginas/tiempo
            pages_since_rotate += 1
            now = time.time()
            if pages_since_rotate >= pages_limit or (now - last_rotate_ts) >= time_limit_sec:
                # Señalización inmediata de rotación y heartbeats para no quedar "mudo"
                NEED_RESTART = True
                log_sys("[STAGE3] Umbral de rotación alcanzado; solicitando reinicio (inmediato)")
                log_sys("[ROTATE] cerrando ciclo actual…")
                save_persistent_state(state)
                # Señal al productor para cortar de inmediato y permitir join() rápido
                try:
                    stop_event.set()
                except Exception:
                    pass
                log_sys("[ROTATE] estado persistido; saliendo a bucle de re‑arranque")
                break

        # Guardar cookies al salir solo si no venimos de WAF alto
        if not (state.get('last_waf_high') is True):
            try:
                save_cookies(page, COOKIE_STORE_PATH)
            except Exception:
                pass


def save_persistent_state(state: dict, path: str = STATE_FILE):
    """Guarda el estado persistente tras cada página o hito importante."""
    try:
        state['fecha_ultimo_proceso'] = time.time()
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2, ensure_ascii=False)
        if SHOW_STATE_SAVES:
            log_sys(f"Estado guardado en {path}")
    except Exception as e:
        log_warn(f"No se pudo guardar estado en {path}: {e}")


def _extract_ids_from_current_listado(page):
    """Extrae todos los (id_anuncio, url) de la página de listado actual, preservando el orden visual.

    Retorna lista de tuplas [(id, url), ...].
    """
    out = []
    try:
        anchors = page.query_selector_all(
            "#main-content > section.items-container.items-list > article > div.item-info-container > a"
        )
        for a in anchors:
            try:
                href = a.get_attribute('href')
            except Exception:
                href = None
            if not href:
                continue
            abs_url = urljoin(page.url, href)
            lid = extract_id_from_url(abs_url)
            if lid:
                out.append((lid, abs_url))
    except Exception as e:
        log_warn(f"No se pudieron extraer IDs del listado: {type(e).__name__}: {e}")
    return out


def _goto_listado_desc(page):
    """Navega a la página de listado ordenado por recientes (desc)."""
    status = navigate_and_stabilize(page, WEB_LISTADO_DESC, tag='listado_desc', wait_networkidle=False)
    risk = classify_waf_risk(status)
    if risk in ("MUY ALTO", "ALTO"):
        log_waf(f"{risk} en listado_desc")
        try:
            handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
        except Exception:
            pass
    gentle_scroll(page)
    wait_ms = random.randint(1200, 2200)
    log_sys(f"[RHYTHM] listado_desc pausa ~{wait_ms}ms")
    safe_wait_page(page, wait_ms)


def _goto_listado_asc(page):
    """Navega a la página de listado ordenado por antiguos (asc)."""
    status = navigate_and_stabilize(page, WEB_LISTADO_ASC, tag='listado_asc', wait_networkidle=False)
    risk = classify_waf_risk(status)
    if risk in ("MUY ALTO", "ALTO"):
        log_waf(f"{risk} en listado_asc")
        try:
            handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
        except Exception:
            pass
    gentle_scroll(page)
    safe_wait_page(page, random.randint(120, 250))


def _collect_page_urls_forward(page, max_pages=10):
    """Colecciona URLs de páginas consecutivas hacia adelante empezando en la actual.

    Uso: para simular recorrido ascendente procesando luego en orden inverso.
    """
    urls = [page.url]
    current = 1
    while current < max_pages:
        next_url = get_next_page_url(page)
        if not next_url:
            break
        status = navigate_and_stabilize(page, next_url, tag=f'listado_p{current+1}', wait_networkidle=False)
        risk = classify_waf_risk(status)
        if risk in ("MUY ALTO", "ALTO"):
            log_waf(f"{risk} navegando páginas forward")
            break
        urls.append(page.url)
        gentle_scroll(page)
        wait_ms = random.randint(1600, 3200)
        log_sys(f"[RHYTHM] forward pausa ~{wait_ms}ms")
        safe_wait_page(page, wait_ms)
        current += 1
    return urls


def phase1_captura_nuevos(page, state: dict, consecutivos_stop: int = 50, max_pages: int = 20):
    """FASE 1 — Captura de nuevos (más recientes primero).

    - Orden: descendente por fecha (recientes primero).
    - Reglas:
        * contador_consecutivos_conocidos ++ cuando vemos un conocido.
        * Si encontramos un nuevo, lo PROCESAMOS (scrape) y lo añadimos a ids_conocidos, y reseteamos contador.
        * Si contador >= consecutivos_stop: guardar ese id como ultimo_pivote_fase1 y terminar Fase 1.
    - Persistimos estado tras cada página.
    """
    global NEED_RESTART
    known = set(state.get('ids_conocidos') or [])
    contador = 0
    _goto_listado_desc(page)

    page_idx = 1
    while page_idx <= max_pages:
        log_sys(f"[F1] Página recientes #{page_idx}")
        ids_urls = _extract_ids_from_current_listado(page)
        if not ids_urls:
            log_warn("[F1] Sin IDs en esta página; deteniendo Fase 1")
            break

        for (lid, lurl) in ids_urls:
            if lid in known:
                contador += 1  # conocido consecutivo
                # ¿hemos alcanzado el umbral de corte?
                if contador >= consecutivos_stop:
                    state['ultimo_pivote_fase1'] = lid  # pivot en el último conocido
                    save_persistent_state(state)
                    log_sys(f"[F1] Umbral alcanzado con {lid}; pivot fijado y Fase 1 terminada")
                    return
                continue

            # NO está en conocidos → procesar (scrape) y agregar
            try:
                status = navigate_and_stabilize(page, lurl, tag='anuncio_f1')
                risk = classify_waf_risk(status)
                if risk in ("MUY ALTO", "ALTO"):
                    log_waf(f"{risk} en anuncio F1 {lid}; marcando NEED_RESTART")
                    try:
                        handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
                    except Exception:
                        pass
                    NEED_RESTART = True
                    save_persistent_state(state)
                    return

                humanize_anuncio_before_parse(page)
                data = parse_anuncio(page)
                if data and data.get('id_anuncio'):
                    xlsx_path = os.path.join(os.path.dirname(__file__), 'scraper.xlsx')
                    append_row_xlsx(xlsx_path, data)
                    known.add(lid)
                    state['ids_conocidos'] = list(known)
                    contador = 0  # reset de racha de conocidos
                # volver al listado
                try:
                    page.go_back(); page.wait_for_load_state('domcontentloaded')
                    safe_wait_page(page, random.randint(60, 140))
                except Exception as e:
                    log_warn(f"Volver a listado (F1) falló: {e}")
                    save_persistent_state(state)
                    return
            except Exception as e:
                log_warn(f"[F1] Error procesando {lid}: {e}")
                # continuar con el siguiente id

        # Persistir tras página
        save_persistent_state(state)

        # Ir a siguiente página de listados recientes
        next_url = get_next_page_url(page)
        if not next_url:
            log_sys("[F1] No hay más páginas recientes; fin Fase 1")
            break
        status = navigate_and_stabilize(page, next_url, tag=f'listado_desc_p{page_idx+1}', wait_networkidle=False)
        risk = classify_waf_risk(status)
        if risk in ("MUY ALTO", "ALTO"):
            log_waf(f"{risk} en navegación F1; NEED_RESTART")
            NEED_RESTART = True
            save_persistent_state(state)
            return
        gentle_scroll(page)
        safe_wait_page(page, random.randint(80, 160))
        page_idx += 1


def phase2_procesa_antiguos(page, state: dict, consecutivos_stop: int = 50, max_pages_window: int = 30):
    """FASE 2 — Procesamiento de antiguos en orden ascendente (más antiguos primero).

    Reglas de parada:
      - encontrar `ultimo_pivote_fase1`,
      - alcanzar `consecutivos_stop` conocidos consecutivos,
      - no hay más páginas,
      - WAF alto.
    Reanudación:
      - Si `ultimo_procesado_fase2` existe, buscarlo y continuar desde ese punto.
    """
    global NEED_RESTART
    known = set(state.get('ids_conocidos') or [])
    pivot = state.get('ultimo_pivote_fase1')
    last_old = state.get('ultimo_procesado_fase2')

    # 1) Ir a listado ASC (antiguos primero)
    _goto_listado_asc(page)

    # 2) Si tenemos ultimo_procesado_fase2, navegar páginas hasta encontrarlo
    page_index = 1
    found_start = False if last_old else True
    while page_index <= max_pages_window and not found_start:
        ids_urls = _extract_ids_from_current_listado(page)
        ids = [lid for (lid, _) in ids_urls]
        if last_old in ids:
            found_start = True
            break
        next_url = get_next_page_url(page)
        if not next_url:
            break
        status = navigate_and_stabilize(page, next_url, tag=f'f2_seek_p{page_index+1}', wait_networkidle=False)
        if classify_waf_risk(status) in ("MUY ALTO", "ALTO"):
            log_waf("[F2] Riesgo WAF buscando punto de reanudación")
            NEED_RESTART = True
            save_persistent_state(state)
            return
        gentle_scroll(page)
        safe_wait_page(page, random.randint(800, 1600))
        page_index += 1

    consecutivos = 0
    processed_pages = 0
    while processed_pages < max_pages_window:
        ids_urls = _extract_ids_from_current_listado(page)
        if not ids_urls:
            log_warn("[F2] Página sin IDs; detener")
            break

        start_idx = 0
        if last_old and not found_start:
            # Si no lo encontramos en ventana, comenzamos desde principio igualmente
            pass
        elif last_old and found_start:
            # Continuar después de last_old
            ids = [lid for (lid, _) in ids_urls]
            if last_old in ids:
                start_idx = ids.index(last_old) + 1

        for (lid, lurl) in ids_urls[start_idx:]:
            if pivot and lid == pivot:
                log_sys(f"[F2] Alcanzado pivot {pivot}; terminar Fase 2")
                save_persistent_state(state)
                return

            if lid in known:
                consecutivos += 1
                if consecutivos >= consecutivos_stop:
                    log_sys(f"[F2] {consecutivos} conocidos consecutivos; terminar")
                    save_persistent_state(state)
                    return
                continue

            try:
                status = navigate_and_stabilize(page, lurl, tag='anuncio_f2')
                # Gestión de WAF F2 con backoff + rotación
                risk = classify_waf_risk(status)
                MUY_ALTO_STATUSES = {403, 429, 418, 444}
                s = None
                try:
                    s = int(status) if status is not None else None
                except Exception:
                    s = None
                if s in MUY_ALTO_STATUSES:
                    # Incrementar contador persistente
                    try:
                        state['waf_consec_muy_alto_f2'] = int(state.get('waf_consec_muy_alto_f2', 0)) + 1
                    except Exception:
                        state['waf_consec_muy_alto_f2'] = 1
                    save_persistent_state(state)
                    # Backoff breve antes de reiniciar/terminar
                    backoff = random.randint(1800, 3200)
                    log_waf(f"HTTP {s} en anuncio F2 {lid}; backoff ~{backoff}ms")
                    safe_wait_page(page, backoff)
                    if state['waf_consec_muy_alto_f2'] >= 2:
                        log_waf("[F2] 2 bloqueos MUY_ALTO consecutivos; detener para análisis")
                        save_persistent_state(state)
                        return
                    else:
                        log_waf("[F2] Primer MUY_ALTO; solicitar reinicio para rotar proxy/fingerprint")
                        NEED_RESTART = True
                        save_persistent_state(state)
                        return
                # Otros riesgos altos mantienen política de reinicio
                if risk in ("MUY ALTO", "ALTO"):
                    log_waf(f"{risk} en anuncio F2 {lid}; NEED_RESTART")
                    NEED_RESTART = True
                    save_persistent_state(state)
                    return

                # Reset de contador si no hay MUY_ALTO
                if state.get('waf_consec_muy_alto_f2'):
                    state['waf_consec_muy_alto_f2'] = 0

                humanize_anuncio_before_parse(page)
                data = parse_anuncio(page)
                if data and data.get('id_anuncio'):
                    xlsx_path = os.path.join(os.path.dirname(__file__), 'scraper.xlsx')
                    append_row_xlsx(xlsx_path, data)
                    known.add(lid)
                    state['ids_conocidos'] = list(known)
                    state['ultimo_procesado_fase2'] = lid
                    consecutivos = 0
                try:
                    page.go_back(); page.wait_for_load_state('domcontentloaded')
                    safe_wait_page(page, random.randint(60, 140))
                except Exception as e:
                    log_warn(f"Volver a listado (F2) falló: {e}")
                    save_persistent_state(state)
                    return
            except Exception as e:
                log_warn(f"[F2] Error procesando {lid}: {e}")

        save_persistent_state(state)

        next_url = get_next_page_url(page)
        if not next_url:
            log_sys("[F2] No hay más páginas ASC; fin Fase 2")
            break
        status = navigate_and_stabilize(page, next_url, tag=f'f2_next', wait_networkidle=False)
        if classify_waf_risk(status) in ("MUY ALTO", "ALTO"):
            log_waf("[F2] Riesgo WAF navegando ASC")
            NEED_RESTART = True
            save_persistent_state(state)
            return
        gentle_scroll(page)
        safe_wait_page(page, random.randint(80, 160))
        processed_pages += 1

    log_sys("[F2] Fin procesamiento ASC")
    save_persistent_state(state)



def run_daily_two_phase(page, state_path: str = STATE_FILE,
                        f1_consecutivos_stop: int = 50,
                        f2_consecutivos_stop: int = 50,
                        f1_max_pages: int = 20,
                        f2_max_pages_window: int = 30):
    """Orquestador diario de 2 fases.

    - Primera ejecución (sin fecha_ultimo_proceso o sin ids_conocidos): solo Fase 1.
    - Siguientes ejecuciones: Fase 1 (captura de nuevos) y luego Fase 2 (backlog antiguos
      hasta alcanzar pivot o racha de conocidos).
    - Persiste estado tras cada fase.
    """
    state = load_persistent_state(state_path)
    first_run = (not state.get('fecha_ultimo_proceso')) or not (state.get('ids_conocidos') or [])

    log_sys("[DAILY] Inicio orquestación 2 fases")
    if first_run:
        log_sys("[DAILY] Primera ejecución detectada → ejecutar SOLO Fase 1")
        phase1_captura_nuevos(page, state, consecutivos_stop=f1_consecutivos_stop, max_pages=f1_max_pages)
        save_persistent_state(state, state_path)
        return

    # Ejecuciones siguientes: Fase 1 y luego Fase 2 (si no hay NEED_RESTART)
    phase1_captura_nuevos(page, state, consecutivos_stop=f1_consecutivos_stop, max_pages=f1_max_pages)
    save_persistent_state(state, state_path)
    if NEED_RESTART:
        log_sys("[DAILY] NEED_RESTART tras Fase 1; salida anticipada")
        return

    phase2_procesa_antiguos(page, state, consecutivos_stop=f2_consecutivos_stop, max_pages_window=f2_max_pages_window)
    save_persistent_state(state, state_path)

def _ensure_session_user_data_dir(base_dir: str, keep_last: int = 5) -> str:
    """Crea un subdirectorio de sesión y aplica retención (quedarse con últimos N).

    Devuelve la ruta del subdirectorio creado.
    """
    os.makedirs(base_dir, exist_ok=True)
    ts = time.strftime('%Y%m%d_%H%M%S')
    rnd = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=4))
    sess_dir = os.path.join(base_dir, f'sess_{ts}_{rnd}')
    os.makedirs(sess_dir, exist_ok=True)
    try:
        # Mantener solo los últimos keep_last subdirectorios
        all_sess = [p for p in glob.glob(os.path.join(base_dir, 'sess_*')) if os.path.isdir(p)]
        all_sess.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        for old in all_sess[keep_last:]:
            try:
                # Borrado superficial del árbol
                for root, dirs, files in os.walk(old, topdown=False):
                    for f in files:
                        try:
                            os.remove(os.path.join(root, f))
                        except Exception:
                            pass
                    for d in dirs:
                        try:
                            os.rmdir(os.path.join(root, d))
                        except Exception:
                            pass
                os.rmdir(old)
            except Exception:
                pass
    except Exception:
        pass
    return sess_dir


def _gen_sessid() -> str:
    """Genera un sessid aleatorio para Oxylabs (C6)."""
    epoch = int(time.time())
    rnd = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=3))
    return f"cascade_{epoch}_{rnd}"


def log_info(msg):
    """Log de información."""
    console.print(f"[INFO] {msg}", style="cyan")


def log_warn(msg):
    """Log de advertencia."""
    console.print(f"[WARN] {msg}", style="yellow")


# === Verbosity flags (modo compacto con colores) ===
SHOW_NAV = True        # Logs [NAV]
SHOW_BURSTS = False     # Logs de micro-bursts
SHOW_STEPS = True      # Logs de pasos finos [STAGE3 Visitando]
SHOW_PROGRESS = False   # Logs de progreso parciales (cada N items)
SHOW_STATE_SAVES = False  # Mostrar "Estado guardado" frecuentes
SHOW_SYS = True        # Logs [SYS] generales (mantener False para solo SUMMARY)


def log_nav(tag, url, status):
    """Log de navegación compacta (solo hitos)."""
    if not SHOW_NAV:
        return
    style = "green" if (status and 200 <= status < 400) else "bright_red"
    console.print(f"[NAV] {tag}: {url} status={status}", style=style)


def log_sys(msg):
    """Logs internos del sistema (no del sitio/WAF)."""
    if not SHOW_SYS:
        return
    console.print(f"[SYS] {msg}", style="dim")


def log_waf(msg):
    """Logs de señales/alertas del WAF (respuestas del sitio o páginas de desafío)."""
    console.print(f"[WAF] {msg}", style="bright_magenta")


def classify_waf_risk(status):
    """Clasifica el nivel de riesgo según 'knowledge/project_description.md'.

    Devuelve una etiqueta: 'MUY ALTO', 'ALTO', 'MEDIO-ALTO', 'MEDIO', 'BAJO' o None.
    """
    if status is None:
        return None
    try:
        s = int(status)
    except Exception:
        return None
    # MUY ALTO: 429, 403, 418, 444
    if s in (429, 403, 418, 444):
        return 'MUY ALTO'
    # ALTO: 520-529, 406, 413, 414
    if (520 <= s <= 529) or s in (406, 413, 414):
        return 'ALTO'
    # MEDIO-ALTO: 301/302 (CAPTCHA/redirección sospechosa)
    if s in (301, 302):
        return 'MEDIO-ALTO'
    # MEDIO: 502, 400, 405
    if s in (502, 400, 405):
        return 'MEDIO'
    # BAJO: 451, 503
    if s in (451, 503):
        return 'BAJO'
    return None


def get_mac_fonts(seed=None, count=10):
    """Devuelve lista de fonts directamente para Camoufox.

    - seed: fija la semilla para reproducibilidad.
    - count: número de fuentes a devolver.
    """
    if seed:
        random.seed(seed)

    custom_fonts = [
        # Microsoft Office
        "Calibri", "Cambria", "Consolas", "Constantia", "Corbel", "Candara",
        # Adobe
        "Myriad Pro", "Minion Pro", "Adobe Garamond Pro", "Adobe Caslon Pro",
        "Proxima Nova", "Brandon Grotesque", "Acumin Pro", "Adobe Arabic",
        # Google Fonts populares
        "Roboto", "Open Sans", "Lato", "Montserrat", "Poppins", "Inter",
        "Raleway", "Ubuntu", "Oswald", "Playfair Display", "Noto Sans",
        "Work Sans", "Nunito", "PT Sans", "Source Sans Pro", "Merriweather",
        # Desarrollo
        "Fira Code", "JetBrains Mono", "Source Code Pro", "Cascadia Code",
        "Hack", "Inconsolata", "Ubuntu Mono", "Droid Sans Mono", "Victor Mono",
        # Diseño/Apps
        "SF Pro Display", "Circular", "Graphik", "Apercu", "Gilroy",
        "Avenir Next Pro", "DIN", "Gotham", "Neue Haas Grotesk Display",
        "Atlas Grotesk", "Maison Neue", "Suisse Int'l", "Akkurat",
        # Misc apps populares
        "Slack-Lato", "WhatsApp Font", "Spotify Circular", "Netflix Sans",
        "YouTube Sans", "IBM Plex Sans", "Samsung Sans", "Google Sans"
    ]

    return random.sample(custom_fonts, min(count, len(custom_fonts)))


def get_webgl_config(seed=None):
    """Devuelve una tupla (vendor, renderer) para Camoufox.webgl_config.

    Variamos por sesión para rotar fingerprint gráfico de forma ligera.
    """
    try:
        if seed is not None:
            random.seed(str(seed))
    except Exception:
        pass
    # Config fijo conocido compatible en macOS para evitar ValueError
    return ("Intel Inc.", "Intel Iris OpenGL Engine")


def safe_wait_page(page, ms):
    """Espera segura gestionando cierre de página/contexto.

    Si la página se cierra, registramos y continuamos sin error duro.
    """
    try:
        if page.is_closed():  # Página ya cerrada
            log_warn("Página cerrada antes de esperar.")
            return
        page.wait_for_timeout(ms)  # Espera en ms
    except Exception as e:  # Maneja TargetClosedError y similares
        log_warn(f"Espera interrumpida: {type(e).__name__}: {e}")


def navigate_and_stabilize(page, url, tag, wait_networkidle=True):
    """Navega a 'url' y espera estados razonables.

    - tag: etiqueta corta para logs (p.ej. 'home', 'mapa').
    - wait_networkidle: False para páginas con red activa (listado).
    """
    try:
        resp = page.goto(url)  # Ir a URL
        status = resp.status if resp else None
        log_nav(tag, page.url, status)
        # Señalizar estados de posible WAF con nivel de riesgo
        risk = classify_waf_risk(status)
        if risk:
            log_waf(f"{risk} · {tag}: {page.url} status={status}")
        page.wait_for_load_state("domcontentloaded")  # DOM listo
        if wait_networkidle:
            page.wait_for_load_state("networkidle")  # Red estable
        else:
            page.wait_for_timeout(random.randint(120, 300))  # Jitter corto
        return status
    except Exception as e:
        log_warn(f"Navegación fallida [{tag}]: {type(e).__name__}: {e}")
        return None


def gentle_scroll(page):
    """Realiza un scroll suave para gesto humano y carga perezosa."""
    try:
        page.evaluate("window.scrollBy(0, 600)")  # Scroll corto
        page.wait_for_timeout(random.randint(80, 150))  # Jitter corto
        page.evaluate("window.scrollBy(0, 1200)")  # Otro scroll
        page.wait_for_timeout(random.randint(80, 150))
        page.evaluate("window.scrollTo(0, 0)")  # Volver arriba
    except Exception as e:
        log_warn(f"Scroll interrumpido: {type(e).__name__}: {e}")


def humanize_listado(page):
    """Pequeñas interacciones en web_listado (C2): micro-scrolls aleatorios."""
    try:
        # Scrolls aleatorios cortos
        steps = random.randint(1, 2)
        for _ in range(steps):
            dy = random.randint(300, 900)
            page.evaluate(f"window.scrollBy(0, {dy})")
            page.wait_for_timeout(random.randint(60, 180))
        # Vuelve arriba a veces
        if random.random() < 0.4:
            page.evaluate("window.scrollTo(0, 0)")
            page.wait_for_timeout(random.randint(50, 120))
    except Exception as e:
        log_warn(f"Interacción listado interrumpida: {type(e).__name__}: {e}")


def humanize_anuncio_before_parse(page):
    """Interacciones en web_anuncio antes de parsear (C2)."""
    try:
        # 2-3 scrolls con pausas aleatorias
        n = random.randint(2, 3)
        for _ in range(n):
            dy = random.randint(500, 1400)
            page.evaluate(f"window.scrollBy(0, {dy})")
            page.wait_for_timeout(random.randint(80, 180))
        # Subir un poco y estabilizar
        if random.random() < 0.5:
            page.evaluate("window.scrollBy(0, -400)")
            page.wait_for_timeout(random.randint(60, 140))
    except Exception as e:
        log_warn(f"Interacción anuncio interrumpida: {type(e).__name__}: {e}")


def _safe_text(el):
    """Devuelve el innerText limpio de un elemento o None."""
    try:
        if not el:
            return None
        txt = el.inner_text()
        return (txt or '').strip()
    except Exception:
        return None


def _q_text(page, sel):
    """Query single selector y texto limpio."""
    try:
        el = page.query_selector(sel)
        return _safe_text(el)
    except Exception:
        return None


def _q_all_texts(page, sel):
    """Query all selector y devuelve lista de textos no vacíos."""
    out = []
    try:
        els = page.query_selector_all(sel)
        for el in els:
            t = _safe_text(el)
            if t:
                out.append(t)
    except Exception:
        pass
    return out


def parse_anuncio(page):
    """Extrae campos definidos en knowledge/project_description.md de una web_anuncio.

    Devuelve un dict con: id_anuncio, precio, m2, habitaciones, piso__exterior_ascensor, detalles,
    caracteristicas_basicas (lista), edificio (lista), equipamiento (lista), certificado_energetico (lista).
    """
    url = page.url or ''
    # id_anuncio desde URL (último segmento numérico)
    id_anuncio = None
    try:
        parts = [p for p in url.split('/') if p]
        for p in parts[::-1]:
            if p.isdigit():
                id_anuncio = p
                break
    except Exception:
        id_anuncio = None

    # Selectores de Datos únicos
    sel_precio = (
        "#main > div > main > section.detail-info.ide-box-detail-first-picture.ide-box-detail--reset.overlay-box > "
        "section > div.info-data > span > span"
    )
    sel_m2 = (
        "#main > div > main > section.detail-info.ide-box-detail-first-picture.ide-box-detail--reset.overlay-box > "
        "section > div.info-features > span:nth-child(1)"
    )
    sel_habitaciones = (
        "#main > div > main > section.detail-info.ide-box-detail-first-picture.ide-box-detail--reset.overlay-box > "
        "section > div.info-features > span:nth-child(2)"
    )
    sel_piso_ext_asc = (
        "#main > div > main > section.detail-info.ide-box-detail-first-picture.ide-box-detail--reset.overlay-box > "
        "section > div.info-features > span:nth-child(3)"
    )
    sel_detalles = (
        "#main > div > main > section.detail-info.ide-box-detail-first-picture.ide-box-detail--reset.overlay-box > "
        "section > div.info-features"
    )
    # Nuevos campos (Datos únicos)
    sel_comentario_anunciante = (
        "#main > div > main > section.detail-info.ide-box-detail-first-picture.ide-box-detail--reset.overlay-box > "
        "section > div.commentsContainer > div.comment > div > p"
    )
    sel_ubicacion = "#headerMap > h2"

    precio = _q_text(page, sel_precio)
    m2 = _q_text(page, sel_m2)
    habitaciones = _q_text(page, sel_habitaciones)
    piso_ext_asc = _q_text(page, sel_piso_ext_asc)
    detalles = _q_text(page, sel_detalles)
    comentario_anunciante = _q_text(page, sel_comentario_anunciante)
    ubicacion = _q_text(page, sel_ubicacion)

    # Selectores de Datos listados (listas de textos)
    sel_caract = "#details > div.details-property > div.details-property-feature-one > div > ul li"
    sel_edificio = "#details > div.details-property > div.details-property-feature-one > div:nth-child(4) > ul li"
    sel_equip = "#details > div.details-property > div.details-property-feature-two > div:nth-child(2) > ul li"
    sel_cert = "#details > div.details-property > div.details-property-feature-two > div:nth-child(4) > ul li"

    caracteristicas_basicas = _q_all_texts(page, sel_caract)
    edificio = _q_all_texts(page, sel_edificio)
    equipamiento = _q_all_texts(page, sel_equip)
    certificado_energetico = _q_all_texts(page, sel_cert)

    return {
        "id_anuncio": id_anuncio,
        "precio": precio,
        "m2": m2,
        "habitaciones": habitaciones,
        "piso_exterior_ascensor": piso_ext_asc,
        "detalles": detalles,
        "comentario_anunciante": comentario_anunciante,
        "ubicación": ubicacion,
        "caracteristicas_basicas": caracteristicas_basicas,
        "edificio": edificio,
        "equipamiento": equipamiento,
        "certificado_energetico": certificado_energetico,
        "url": url,
    }


# === FUNCIONES DE PERSISTENCIA DE IDS ===

def load_known_ids(ids_file="ids_seen.json", xlsx_file="scraper.xlsx"):
    """Carga IDs conocidos desde archivo JSON y opcionalmente desde Excel."""
    known_ids = set()
    
    # Cargar desde JSON (rápido)
    if os.path.exists(ids_file):
        try:
            with open(ids_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                known_ids.update(data.get('ids', []))
                log_sys(f"Cargados {len(known_ids)} IDs desde {ids_file}")
        except Exception as e:
            log_warn(f"Error cargando {ids_file}: {e}")
    
    # Cargar desde Excel como backup
    if os.path.exists(xlsx_file):
        try:
            wb = load_workbook(xlsx_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0]:  # id_anuncio en primera columna
                    known_ids.add(str(row[0]))
            log_sys(f"Total IDs únicos conocidos: {len(known_ids)}")
        except Exception as e:
            log_warn(f"Error leyendo IDs desde {xlsx_file}: {e}")
    
    return known_ids


def save_known_ids(known_ids, ids_file="ids_seen.json"):
    """Guarda IDs conocidos en archivo JSON."""
    try:
        data = {'ids': list(known_ids), 'updated': time.time()}
        with open(ids_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        log_sys(f"Guardados {len(known_ids)} IDs en {ids_file}")
    except Exception as e:
        log_warn(f"Error guardando {ids_file}: {e}")


def extract_id_from_url(url):
    """Extrae id_anuncio de una URL de *********."""
    # Formato típico: /inmueble/12345678/
    try:
        parts = url.split('/')
        for i, part in enumerate(parts):
            if part == 'inmueble' and i + 1 < len(parts):
                return parts[i + 1]
    except Exception:
        pass
    return None


# === PERSISTENCIA DE LINKS RECOLECTADOS ===

def save_links_to_visit(links, path="links_to_visit.json"):
    """Guarda la lista de links recolectados en disco para tolerancia a fallos.

    links: lista de tuplas (id_anuncio, url)
    """
    try:
        payload = {
            "links": [{"id": lid, "url": lurl} for (lid, lurl) in links],
            "updated": time.time(),
        }
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2)
        log_sys(f"Persistidos {len(links)} links en {path}")
    except Exception as e:
        log_warn(f"No se pudieron guardar links en {path}: {e}")


def load_links_to_visit(path="links_to_visit.json", known_ids=None):
    """Carga links pendientes desde disco y filtra por IDs ya conocidos.

    Devuelve lista de tuplas (id_anuncio, url).
    """
    global NEED_RESTART
    global NEED_RESTART
    if known_ids is None:
        known_ids = set()
    if not os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        raw = data.get("links", []) or []
        out = []
        for item in raw:
            try:
                lid = str(item.get("id")) if item.get("id") is not None else None
                lurl = item.get("url")
            except Exception:
                lid, lurl = None, None
            if lid and lurl and lid not in known_ids:
                out.append((lid, lurl))
        log_sys(f"Cargados {len(out)} links pendientes desde {path}")
        return out
    except Exception as e:
        log_warn(f"No se pudieron cargar links desde {path}: {e}")
        return []


def append_row_xlsx(filepath, data):
    """Escribe 'data' en Excel. Si no existe, crea el libro y cabecera.

    Sin deduplicación en Fase 2.
    """
    headers = [
        "id_anuncio",
        "url",
        "precio",
        "m2",
        "habitaciones",
        "piso_exterior_ascensor",
        "detalles",
        "comentario_anunciante",
        "ubicación",
        "caracteristicas_basicas",
        "edificio",
        "equipamiento",
        "certificado_energetico",
    ]

    # Serializar listas como texto separadas por " | " para Excel
    row = []
    for h in headers:
        v = data.get(h)
        if isinstance(v, list):
            row.append(" | ".join(v))
        else:
            row.append(v)

    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "anuncios"
        ws.append(headers)
        ws.append(row)
        wb.save(filepath)
        return
    # Append
    wb = load_workbook(filepath)
    ws = wb.active
    # Si el archivo existe pero no tiene cabecera, la añadimos
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(headers)
    ws.append(row)
    wb.save(filepath)


# === FUNCIONES DE RECOLECCIÓN DE LINKS ===

def collect_listado_ids(page, max_pages=5, stop_on_known=True, known_ids=None, save_every_page_path=None):
    """Recolecta id_anuncio únicos de páginas de listado.
    
    Args:
        page: Página de Playwright
        max_pages: Máximo de páginas de listado a procesar
        stop_on_known: Si True, para al encontrar el primer ID conocido
        known_ids: Set de IDs ya conocidos
    
    Returns:
        list: Lista de tuplas (id_anuncio, url_completa) ordenadas
    """
    if known_ids is None:
        known_ids = set()
    
    collected_links = []
    current_page = 1
    
    log_sys(f"Iniciando recolección de links (max_pages={max_pages}, stop_on_known={stop_on_known})")
    
    while current_page <= max_pages:
        log_sys(f"Procesando página {current_page}/{max_pages}")
        
        # Esperar a que cargue el listado
        try:
            page.wait_for_selector(
                "#main-content > section.items-container.items-list",
                timeout=10000,
                state="attached",
            )
        except Exception as e:
            log_warn(f"Error esperando listado en página {current_page}: {e}")
            break
        
        # Recolectar anchors de anuncios
        anchors = page.query_selector_all(
            "#main-content > section.items-container.items-list > "
            "article > div.item-info-container > a"
        )
        
        page_links = []
        for anchor in anchors:
            try:
                href = anchor.get_attribute("href")
                if href:
                    full_url = urljoin(page.url, href)
                    id_anuncio = extract_id_from_url(full_url)
                    
                    if id_anuncio:
                        # Verificar si ya lo conocemos
                        if stop_on_known and id_anuncio in known_ids:
                            log_sys(f"Encontrado ID conocido {id_anuncio}, deteniendo recolección")
                            return collected_links
                        
                        # Verificar si ya lo recolectamos en esta sesión
                        if not any(link[0] == id_anuncio for link in collected_links):
                            page_links.append((id_anuncio, full_url))
                            
            except Exception as e:
                log_warn(f"Error procesando anchor: {e}")
                continue
        
        collected_links.extend(page_links)
        log_sys(f"Recolectados {len(page_links)} links únicos de página {current_page}")
        # Guardado incremental tras cada página para tolerancia a fallos
        try:
            if save_every_page_path:
                save_links_to_visit(collected_links, path=save_every_page_path)
        except Exception as e:
            log_warn(f"No se pudo persistir tras página {current_page}: {e}")
        
        # Intentar ir a siguiente página
        if current_page < max_pages:
            next_url = get_next_page_url(page)
            if not next_url:
                log_sys("No hay más páginas disponibles")
                break
            
            # Navegar a siguiente página
            status = navigate_and_stabilize(
                page, next_url, tag=f"listado_p{current_page + 1}", 
                wait_networkidle=False
            )
            
            # Verificar WAF
            risk = classify_waf_risk(status)
            if risk in ("MUY ALTO", "ALTO"):
                log_waf(f"{risk} durante recolección en página {current_page + 1}")
                NEED_RESTART = True
                return collected_links
            
            if detect_challenge(page):
                log_waf("Desafío detectado durante recolección")
                handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
            
            gentle_scroll(page)
            safe_wait_page(page, random.randint(80, 160))
        
        current_page += 1
    
    log_sys(f"Recolección completada: {len(collected_links)} links únicos")
    return collected_links


def visit_links_in_order(page, links_to_visit, known_ids=None):
    """Visita una lista de links de anuncios en orden secuencial.
    
    Args:
        page: Página de Playwright
        links_to_visit: Lista de tuplas (id_anuncio, url)
        known_ids: Set de IDs conocidos (se actualiza durante el proceso)
    """
    if known_ids is None:
        known_ids = set()
    
    log_sys(f"Iniciando visita secuencial de {len(links_to_visit)} anuncios")
    
    for idx, (id_anuncio, url) in enumerate(links_to_visit, 1):
        log_sys(f"Visitando anuncio {idx}/{len(links_to_visit)}: {id_anuncio}")
        
        # Humanización antes de navegar
        if idx > 1:
            humanize_listado(page)
        
        # Navegar al anuncio
        status = navigate_and_stabilize(page, url, tag=f"anuncio_{idx}")
        
        # Evaluar riesgo WAF
        risk = classify_waf_risk(status)
        if risk in ("MUY ALTO", "ALTO"):
            log_waf(f"{risk} · detener scraper en anuncio {idx}")
            try:
                handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
            except Exception as e:
                log_warn(f"Error manejando desafío: {e}")
            safe_wait_page(page, 2000)
            global NEED_RESTART
            NEED_RESTART = True
            return
        
        # Verificar errores HTTP
        try:
            s = int(status) if status else None
            if s is not None and s >= 400:
                log_warn(f"HTTP {s} en anuncio {idx} · omitiendo")
                continue
        except Exception as e:
            log_warn(f"Error procesando status {status}: {e}")
        
        # Detección de desafíos
        if detect_challenge(page):
            log_waf("Desafío detectado en anuncio")
            handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
        
        # Humanización antes de parsear
        humanize_anuncio_before_parse(page)
        
        # Extraer y guardar datos
        try:
            data = parse_anuncio(page)
            if data and data.get("id_anuncio"):
                xlsx_path = os.path.join(os.path.dirname(__file__), "scraper.xlsx")
                append_row_xlsx(xlsx_path, data)
                
                # Añadir a IDs conocidos
                known_ids.add(id_anuncio)
                
                log_sys(f"Datos guardados para anuncio {idx}/{len(links_to_visit)}")
            else:
                log_warn(f"Datos incompletos para anuncio {idx}")
        except Exception as e:
            log_warn(f"Error extrayendo datos: {type(e).__name__}: {e}")
        
        # Scroll y pausa
        gentle_scroll(page)
        safe_wait_page(page, random.randint(120, 220))
        
        # Cooldown periódico (modo rápido: jitter corto)
        if idx % random.randint(8, 10) == 0:
            pause = random.randint(120, 220)
            log_sys(f"Cooldown breve ~{pause}ms tras anuncio {idx}")
            safe_wait_page(page, pause)


def visit_listado_anuncios(page, max_anuncios):
    """Visita hasta 'max_anuncios' anuncios desde la página de listado actual.

    Usa los selectores definidos en knowledge/project_description.md
    (article:nth-child(2), 4, 6, ...), extrae el href del <a> y navega.
    """
    global NEED_RESTART
    try:
        # 1) Intento principal: seleccionar todos los anchors del listado en una sola pasada
        try:
            page.wait_for_selector(
                "#main-content > section.items-container.items-list",
                timeout=2000,
                state="attached",
            )
        except Exception:
            pass

        anchors = page.query_selector_all(
            "#main-content > section.items-container.items-list > "
            "article > div.item-info-container > a"
        )

        hrefs = []
        for a in anchors:
            try:
                href = a.get_attribute("href")
            except Exception:
                href = None
            if href:
                hrefs.append(urljoin(page.url, href))

        # deduplicar preservando orden
        seen = set()
        unique_hrefs = []
        for h in hrefs:
            if h not in seen:
                unique_hrefs.append(h)
                seen.add(h)

        targets = unique_hrefs[: max_anuncios]

        # Fallback: si no encontramos anchors, probamos el escaneo secuencial 1..N
        if not targets:
            found = 0
            attempts = 0
            nth = 1  # iteración secuencial: 1..N
            MAX_ATTEMPTS = 30  # ventana máxima por listado
            while found < max_anuncios and attempts < MAX_ATTEMPTS:
                sel = (
                    f"#main-content > section.items-container.items-list > "
                    f"article:nth-child({nth}) > div.item-info-container > a"
                )
                appeared = False
                try:
                    page.wait_for_selector(sel, timeout=1000, state="attached")
                    appeared = True
                except Exception:
                    appeared = False
                link = page.query_selector(sel) if appeared else None
                if link:
                    href = link.get_attribute("href")
                    if href:
                        targets.append(urljoin(page.url, href))
                        found += 1
                nth += 1
                attempts += 1

        # Navegar a los destinos recolectados
        idx = 0
        for abs_url in targets:
            idx += 1
            # C1: pequeña humanización en listado antes de abrir siguiente
            humanize_listado(page)
            status = navigate_and_stabilize(page, abs_url, tag=f"anuncio_{idx}")
            # Evaluar riesgo y actuar
            risk = classify_waf_risk(status)
            if risk in ("MUY ALTO", "ALTO"):
                log_waf(f"{risk} · detener scraper en anuncio_{idx}")
                # Intento de manejo visible si procede
                try:
                    handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
                except Exception:
                    pass
                # Salida inmediata sin cooldown
                NEED_RESTART = True
                log_sys("[ROTATE] salida inmediata tras posible desafío; re‑arrancando")
                return
            if status is not None:
                try:
                    s = int(status)
                except Exception:
                    s = None
                # Si hay error HTTP, no parsear ni escribir
                if s is not None and s >= 400:
                    log_warn(f"HTTP {s} en anuncio_{idx} · se omite extracción y se vuelve al listado")
                    try:
                        page.go_back()
                        page.wait_for_load_state("domcontentloaded")
                        page.wait_for_timeout(random.randint(20, 80))
                    except Exception as e:
                        log_warn(f"Volver a listado tras error falló: {type(e).__name__}: {e}")
                        break
                    continue
            if detect_challenge(page):
                log_waf("MEDIO-ALTO · desafío detectado en web_anuncio")
                handle_challenge_if_any(page.url, COOKIE_STORE_PATH)
            # C2: interacciones antes de parsear
            humanize_anuncio_before_parse(page)
            # Extraer datos y volcar a Excel antes del scroll
            try:
                data = parse_anuncio(page)
                xlsx_path = os.path.join(os.path.dirname(__file__), "scraper.xlsx")
                append_row_xlsx(xlsx_path, data)
                log_sys(f"Datos escritos para anuncio_{idx} -> {xlsx_path}")
            except Exception as e:
                log_warn(f"No se pudo extraer/escribir datos: {type(e).__name__}: {e}")
            gentle_scroll(page)  # micro-scroll solo en web_anuncio
            # C1: dwell aleatorio entre anuncios (modo rápido)
            safe_wait_page(page, random.randint(120, 220))
            # C1: cooldown periódico cada ~8-10 anuncios (modo rápido)
            if idx % random.randint(8, 10) == 0:
                pause = random.randint(120, 220)
                log_sys(f"Cooldown breve ~{pause}ms tras anuncio_{idx}")
                safe_wait_page(page, pause)
            # volver al listado
            try:
                page.go_back()
                page.wait_for_load_state("domcontentloaded")
                page.wait_for_timeout(random.randint(120, 180))
            except Exception as e:
                log_warn(f"Volver a listado falló: {type(e).__name__}: {e}")
                break
    except Exception as e:
        log_warn(f"Visita de anuncios interrumpida: {type(e).__name__}: {e}")


def get_next_page_url(page):
    """Devuelve la URL absoluta de 'next_page' si existe en la página actual."""
    try:
        sel = (
            "#main-content > section.items-container.items-list > div.pagination > "
            "ul > li.next > a"
        )
        a = page.query_selector(sel)
        if not a:
            return None
        href = a.get_attribute("href")
        if not href:
            return None
        return urljoin(page.url, href)
    except Exception:
        return None


def detect_challenge(page):
    """Heurística simple para detectar CAPTCHA/Turnstile/Access denied.

    Busca elementos comunes en *********/CF: 'cf-challenge', 'turnstile', 'g-recaptcha',
    textos de 'are you human' o 'access denied'. Devuelve True si parece un desafío.
    """
    try:
        return page.evaluate(
            """
            () => {
              const txt = (s) => (s||'').toLowerCase();
              const hay = (sel) => !!document.querySelector(sel);
              const body = document.body ? txt(document.body.innerText) : '';
              if (hay('[id*="cf-challenge"], [class*="cf-challenge"], iframe[src*="challenge"], .cf-turnstile')) return true;
              if (hay('[class*="turnstile"], [id*="turnstile"], .g-recaptcha, #cf-error-details')) return true;
              if (body.includes('are you human') || body.includes('access denied') || body.includes('captcha')) return true;
              return false;
            }
            """
        )
    except Exception:
        return False


def handle_challenge_if_any(current_url, cookie_store_path: str):
    """Si hay un desafío, abre ventana visible sin contexto persistente.

    Flujo:
      - Abrir Camoufox headful SIN persistent_context.
      - Inyectar cookies guardadas (si existen).
      - Navegar a la URL del desafío y esperar hasta 180s a que desaparezca.
      - Guardar cookies al final para mantener validación.
    """
    try:
        log_warn("Desafío detectado. Abriendo ventana visible para resolver...")
        with Camoufox(
            headless=False,
            main_world_eval=False,
            persistent_context=False,
            enable_cache=False,
            block_webrtc=True,
            block_webgl=True,
            block_images=False,  # Necesario para ver el desafío
            os="macos",
            fonts=get_mac_fonts(SESSID, 12),
            geoip=True,
            humanize=True,
        ) as visible_browser:
            vpage = visible_browser.new_page()
            vpage.set_default_navigation_timeout(90000)
            vpage.set_default_timeout(90000)
            # Inyectar cookies antes de navegar
            try_inject_cookies(vpage, cookie_store_path)
            resp = vpage.goto(current_url)
            status = resp.status if resp else None
            log_nav("challenge_headful", vpage.url, status)
            # Esperar hasta 180s a que el desafío desaparezca
            total = 0
            while total < 180000:
                vpage.wait_for_timeout(2000)
                total += 2000
                try:
                    if not detect_challenge(vpage):
                        log_info("Desafío resuelto. Continuando...")
                        break
                except Exception:
                    pass
            # Guardar cookies del contexto efímero
            try:
                save_cookies(vpage, cookie_store_path)
            except Exception as e:
                log_warn(f"No se pudieron guardar cookies tras challenge: {e}")
    except Exception as e:
        log_warn(f"No se pudo abrir ventana visible: {type(e).__name__}: {e}")


# Variables de sesión/proxy para la ejecución actual
NEED_RESTART = False  # Señal opcional para reinicio tras WAF alto (no usado aún)
SESSION_USER_DATA_DIR = _ensure_session_user_data_dir(USER_DATA_DIR_BASE, keep_last=5)

# Sticky (para Fase 1: recolección)
STICKY_PROXY_SERVER = 'https://es-pr.oxylabs.io:10001'
PROXY_USERNAME = 'customer-gcbmasd_residencial_xD9F5'
PROXY_PASSWORD = 'HMn6dUTg_WmC6tL'

# Rotating (para Fase 2: visita) — no usado actualmente, mantenido por compatibilidad
ROTATING_PROXY_SERVER = 'https://es-pr.oxylabs.io:10001'

SESSID = _gen_sessid()


def run_once(session_dir: str, sessid: str):
    """Ejecuta un ciclo completo con arquitectura de 3 etapas y pipeline 2↔3.

    - Siempre sticky session
    - Rotación periódica por tiempo/páginas y forzada por NEED_RESTART
    - Persistencia en state.json
    """
    global NEED_RESTART, SESSION_USER_DATA_DIR, SESSID
    NEED_RESTART = False
    SESSION_USER_DATA_DIR = session_dir
    SESSID = sessid

    # Estado y IDs
    state = load_persistent_state()
    known_ids = load_known_ids()
    log_sys(f"[ENTRY] Arranque con {len(known_ids)} IDs conocidos · siguiente={state.get('siguiente_ejecucion')}")
    # Snapshot para resumen por ejecución
    start_counts = (state.get('counts') or {'recientes': 0, 'antiguos': 0, 'total': 0}).copy()
    start_known = len(known_ids)

    # ETAPA 1: Warm-up
    initial_warm_up(SESSID)
    if NEED_RESTART:
        return

    # ETAPAS 2 y 3: pipeline productor-consumidor
    q = Queue(maxsize=200)
    stop_event = Event()

    # Reinsertar reintentos listos antes de arrancar
    _drain_ready_retries_to_queue(state, q, known_ids)

    # Umbrales de rotación (consumidor los aplicará)
    # Modo eficiente: rotación periódica por tiempo/páginas.
    # Modo seguro (last_waf_high=True): solo rotación forzada por WAF (desactivar periódica).
    if bool(state.get('last_waf_high')):
        time_limit_sec, pages_limit = (10**9, 10**9)
        if SHOW_NAV:
            log_sys("[ROTATE] Safe mode activo: rotación periódica desactivada; solo forzada por WAF")
    else:
        time_limit_sec, pages_limit = _init_rotation_limits()
        if SHOW_NAV:
            log_sys(f"[ROTATE] Eficiente: rotación periódica cada ~{time_limit_sec//60}m o {pages_limit} páginas (con jitter)")

    # Ejecutar ETAPA 2 de forma SECUENCIAL (sin consumidor en paralelo)
    stage2_producer(SESSID, q, stop_event, state, known_ids)

    # Drenar la cola del productor a memoria para ETAPA 3
    collected = []
    try:
        while True:
            item = q.get_nowait()
            if item is None:
                break
            collected.append(item)
    except Empty:
        pass

    # Persistir una versión mínima de los enlaces (tolerancia a fallos)
    try:
        pairs = []
        for it in collected:
            try:
                if isinstance(it, (tuple, list)) and len(it) >= 2:
                    pairs.append((it[0], it[1]))
            except Exception:
                continue
        if pairs:
            save_links_to_visit(pairs)
    except Exception:
        pass

    # Preparar ETAPA 3 de forma SECUENCIAL, reinyectando en una nueva cola
    q2 = Queue(maxsize=200)
    stop_event2 = Event()
    for it in collected:
        try:
            q2.put_nowait(it)
        except Exception:
            continue
    try:
        q2.put_nowait(None)
    except Exception:
        pass

    stage3_consumer(SESSID, q2, stop_event2, state, known_ids, time_limit_sec, pages_limit)

    # Persistencia final
    save_persistent_state(state)
    save_known_ids(known_ids)
    # Resumen por ejecución
    try:
        end_counts = state.get('counts') or {'recientes': 0, 'antiguos': 0, 'total': 0}
        delta_recientes = int(end_counts.get('recientes', 0)) - int(start_counts.get('recientes', 0))
        delta_antiguos = int(end_counts.get('antiguos', 0)) - int(start_counts.get('antiguos', 0))
        delta_total_items = int(end_counts.get('total', 0)) - int(start_counts.get('total', 0))
        end_known = len(known_ids)
        delta_known = end_known - start_known
        console.print(
            f"[SUMMARY] Ejecución: +R={delta_recientes} +A={delta_antiguos} +T={delta_total_items} · "
            f"Acumulado: R={end_counts.get('recientes',0)} A={end_counts.get('antiguos',0)} T={end_counts.get('total',0)} · "
            f"IDs únicos totales={end_known} (Δ={delta_known})",
            style="bold white"
        )
    except Exception as e:
        log_warn(f"No se pudo imprimir SUMMARY: {e}")
    log_sys("[ENTRY] run_once finalizado")
    return

    # (Bloque antiguo de FASE 1/2 fue removido; reemplazado por pipeline de ETAPAS 2 y 3)

# Bucle de auto-reinicio C4+C6: regenerar sesión y sessid cuando WAF alto o proxy inválido
restart_backoff = 0  # cooldown incremental ante WAF alto
while True:
    session_dir = _ensure_session_user_data_dir(USER_DATA_DIR_BASE, keep_last=5)
    sessid = _gen_sessid()
    log_sys(f"Iniciando run con session_dir={session_dir} sessid={sessid}")
    try:
        run_once(session_dir, sessid)
    except Exception as e:
        # Proxy inválido (p.ej. 407) o error al inicializar Camoufox/ipify
        msg = f"{type(e).__name__}: {e}"
        if "InvalidProxy" in msg or "Proxy" in msg or "407" in msg:
            log_waf("Proxy inválido/407 al arrancar. Regenerando sesión y sessid…")
            continue
        # Playwright Sync API dentro de un loop asyncio: tratar como recuperable
        if "Playwright Sync API inside the asyncio loop" in msg:
            log_warn("Playwright detectó loop asyncio; aplicando cooldown y reintentando con nueva sesión…")
            try:
                # Cooldown breve fijo para no martillear
                time.sleep(4)
            except Exception:
                pass
            continue
        # Otros fallos: mostrar y salir para inspección
        log_warn(f"Fallo no controlado en run_once: {msg}")
        break
    if NEED_RESTART:
        # Aumentar cooldown incremental si seguimos bajo WAF alto
        try:
            st = load_persistent_state()
            waf_high = bool(st.get('last_waf_high'))
        except Exception:
            waf_high = False
        if waf_high:
            restart_backoff = min(60, (restart_backoff or 4) * 2)  # 4→8→16→32→60
            jitter = random.randint(-restart_backoff//4, restart_backoff//4)
            wait_s = max(3, restart_backoff + jitter)
            log_waf(f"Cooldown antes de reintentar warm-up: ~{wait_s}s (anti-hammer)")
            try:
                time.sleep(wait_s)
            except Exception:
                pass
        else:
            restart_backoff = 0
        log_waf("Reinicio solicitado por WAF alto. Regenerando sesión y proxy...")
        continue
    else:
        log_sys("Finalizado sin necesidad de reinicio.")
        break
